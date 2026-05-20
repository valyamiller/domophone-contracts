from datetime import datetime, timedelta
from flask import send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os
import io

class ReportGenerator:
    
    @staticmethod
    def generate_expiring_contracts_report(clients):
        """Генерация отчёта по истекающим договорам"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Истекающие договоры"
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовки
        headers = ['№ договора', 'Лицевой счёт', 'ФИО', 'Адрес', 'Телефон', 
                   'Дата начала', 'Дата окончания', 'Дней до окончания', 'Статус', 'Оформил']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Данные
        row = 2
        for client in clients:
            days = client.days_until_expiry()
            if days < 0:
                status = "Истёк"
            elif days <= 7:
                status = "Срочно (менее 7 дней)"
            elif days <= 14:
                status = "Осталось 14 дней"
            elif days <= 30:
                status = "Осталось 30 дней"
            else:
                status = "Активен"
            
            ws.cell(row=row, column=1, value=client.contract_number).border = border
            ws.cell(row=row, column=2, value=client.personal_account).border = border
            ws.cell(row=row, column=3, value=client.full_name).border = border
            ws.cell(row=row, column=4, value=client.full_address).border = border
            ws.cell(row=row, column=5, value=client.phone).border = border
            ws.cell(row=row, column=6, value=client.contract_start.strftime('%d.%m.%Y')).border = border
            ws.cell(row=row, column=7, value=client.contract_end.strftime('%d.%m.%Y')).border = border
            ws.cell(row=row, column=8, value=days if days > 0 else 0).border = border
            ws.cell(row=row, column=9, value=status).border = border
            ws.cell(row=row, column=10, value=client.creator.full_name if client.creator else '—').border = border
            row += 1
        
        # Автоширина колонок
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Сохраняем в байтовый поток
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def generate_monthly_statistics(clients):
        """Генерация месячной статистики"""
        wb = openpyxl.Workbook()
        
        # Лист с общей статистикой
        ws_stats = wb.active
        ws_stats.title = "Общая статистика"
        
        now = datetime.now()
        current_month = now.month
        current_year = now.year
        
        # Подсчёт статистики
        total = len(clients)
        active = sum(1 for c in clients if c.is_active and c.days_until_expiry() > 0)
        expired = sum(1 for c in clients if c.days_until_expiry() is not None and c.days_until_expiry() < 0)
        expiring_30 = sum(1 for c in clients if 0 <= c.days_until_expiry() <= 30)
        expiring_7 = sum(1 for c in clients if 0 <= c.days_until_expiry() <= 7)
        
        # Статистика по месяцам
        contracts_by_month = {}
        for client in clients:
            month = client.contract_start.strftime('%B %Y')
            contracts_by_month[month] = contracts_by_month.get(month, 0) + 1
        
        # Заполняем статистику
        stats_data = [
            ["ПОКАЗАТЕЛЬ", "ЗНАЧЕНИЕ"],
            ["Всего договоров", total],
            ["Активных договоров", active],
            ["Истекших договоров", expired],
            ["Истекает в ближайшие 30 дней", expiring_30],
            ["Истекает в ближайшие 7 дней", expiring_7],
            ["", ""],
            ["ДАТА ФОРМИРОВАНИЯ", datetime.now().strftime('%d.%m.%Y %H:%M')],
        ]
        
        for row, data in enumerate(stats_data, 1):
            for col, value in enumerate(data, 1):
                cell = ws_stats.cell(row=row, column=col, value=value)
                if row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
        
        # Лист с распределением по месяцам
        ws_months = wb.create_sheet("По месяцам")
        ws_months.cell(row=1, column=1, value="Месяц")
        ws_months.cell(row=1, column=2, value="Количество договоров")
        
        row = 2
        for month, count in sorted(contracts_by_month.items()):
            ws_months.cell(row=row, column=1, value=month)
            ws_months.cell(row=row, column=2, value=count)
            row += 1
        
        # Автоширина
        for ws in [ws_stats, ws_months]:
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[col_letter].width = min(max_length + 2, 30)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def generate_clients_export(clients):
        """Экспорт всех клиентов в Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Клиенты"
        
        # Стили
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовки
        headers = ['№ договора', 'Лицевой счёт', 'ФИО', 'Микрорайон', 'Дом', 'Квартира', 
                   'Телефон', 'Дата начала', 'Дата окончания', 'Статус', 'Оформил']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Данные
        row = 2
        for client in clients:
            days = client.days_until_expiry()
            if days is None:
                status = "Неактивен"
            elif days < 0:
                status = "Истёк"
            elif days <= 7:
                status = "Срочно"
            elif days <= 30:
                status = "Скоро"
            else:
                status = "Активен"
            
            ws.cell(row=row, column=1, value=client.contract_number).border = border
            ws.cell(row=row, column=2, value=client.personal_account).border = border
            ws.cell(row=row, column=3, value=client.full_name).border = border
            ws.cell(row=row, column=4, value=client.microdistrict).border = border
            ws.cell(row=row, column=5, value=client.house).border = border
            ws.cell(row=row, column=6, value=client.apartment).border = border
            ws.cell(row=row, column=7, value=client.phone).border = border
            ws.cell(row=row, column=8, value=client.contract_start.strftime('%d.%m.%Y')).border = border
            ws.cell(row=row, column=9, value=client.contract_end.strftime('%d.%m.%Y')).border = border
            ws.cell(row=row, column=10, value=status).border = border
            ws.cell(row=row, column=11, value=client.creator.full_name if client.creator else '—').border = border
            row += 1
        
        # Автоширина
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 30)
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output